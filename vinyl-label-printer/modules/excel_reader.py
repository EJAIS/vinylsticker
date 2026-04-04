"""
Excel reader module.

Loads label records from the "Print" and "Database" sheets of the project
Excel file.  The "Database" sheet is always opened read-only.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

SHEET_PRINT    = "Print"
SHEET_DATABASE = "Database"
COLUMNS        = ("Title", "Artist", "Label", "Country", "Year", "Side")


@dataclass
class LabelRecord:
    title:   str
    artist:  str
    label:   str
    country: str
    year:    str
    side:    str   # always "A" or "B" after normalisation


# ── Public helpers ────────────────────────────────────────────────────────────

def load_print_queue(xlsx_path: Path) -> list[LabelRecord]:
    """Read the 'Print' sheet and return all non-empty rows as LabelRecords.

    The workbook is opened with data_only=True so formula cells resolve to
    their cached value rather than the formula string.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[SHEET_PRINT]
    records = _read_sheet(ws)
    wb.close()
    return records


def load_database(xlsx_path: Path) -> list[LabelRecord]:
    """Read the 'Database' sheet (read-only, never modified).

    Opens the workbook in read_only mode for efficiency.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[SHEET_DATABASE]
    records = _read_sheet(ws)
    wb.close()
    return records


# ── Internal helpers ──────────────────────────────────────────────────────────

def _read_sheet(ws) -> list[LabelRecord]:
    """Iterate rows of *ws*, skip the header row, return LabelRecords."""
    records: list[LabelRecord] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            # Skip header row (Title, Artist, …)
            continue
        record = _row_to_record(row)
        if record is not None:
            records.append(record)
    return records


def _row_to_record(row: tuple) -> LabelRecord | None:
    """Convert an openpyxl row tuple into a LabelRecord.

    Returns None for rows where the Title cell is empty/None so that blank
    rows at the bottom of the sheet are silently ignored.
    """
    # Pad to at least 6 elements in case the row is shorter than expected
    cells = list(row) + [None] * 6
    title, artist, label, country, year, side = cells[:6]

    # Skip rows with no title
    if title is None or str(title).strip() == "":
        return None

    return LabelRecord(
        title=_str(title),
        artist=_str(artist),
        label=_str(label),
        country=_str(country),
        year=_str(year),
        side=_normalise_side(side),
    )


def _str(value: object) -> str:
    """Convert a cell value to a stripped string, defaulting to ''."""
    if value is None:
        return ""
    return str(value).strip()


def _normalise_side(value: object) -> str:
    """Normalise the Side column to a single uppercase letter 'A' or 'B'.

    Handles values like "A", "B", "a", "b", "Seite A", "Side B", etc.
    Defaults to "A" if the value is empty or unrecognisable.
    """
    raw = str(value).strip().upper() if value is not None else ""
    if not raw:
        return "A"
    # Take the last character that is A or B
    for ch in reversed(raw):
        if ch in ("A", "B"):
            return ch
    return "A"
